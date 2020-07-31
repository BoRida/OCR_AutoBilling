import pytesseract #tesseract
import cv2 #computer vision
import time #computation time
import pandas as pd #data frame
import os #file system
import re #regex
import tkinter as tk #GUI
from tkinter import filedialog
from tkinter import *
from openpyxl import Workbook #excel
from PIL import Image

class MainApplication(tk.Frame):

    def spacify_dataframe(self, df):
        data = {'level':[], 'page_num':[], 'block_num':[], 'par_num':[], 'line_num':[], 'word_num':[], 'left':[], 'top':[], 'width':[], 'height':[], 'conf':[], 'text':[]}
        sdf = pd.DataFrame(data)

        index = 0
        sindex = 0
        first_row = {'level': df.loc[df.index[index]].level, 'page_num':df.loc[df.index[index]].page_num, 'block_num':df.loc[df.index[index]].block_num, 'par_num':df.loc[df.index[index]].par_num, 'line_num':df.loc[df.index[index]].line_num, 'word_num':df.loc[df.index[index]].word_num, 'left':df.loc[df.index[index]].left, 'top':df.loc[df.index[index]].top, 'width':df.loc[df.index[index]].width, 'height':df.loc[df.index[index]].height, 'conf':df.loc[df.index[index]].conf, 'text':df.loc[df.index[index]].text}
        sdf = sdf.append(first_row, ignore_index=True)
        index = 1 #first row already added
        sindex = 1
        while index < len(df):

            if sdf.at[sdf.index[sindex-1],'text'].endswith(':') or df.at[df.index[index],'text'].startswith('--'):
                new_row = {'level': df.loc[df.index[index]].level, 'page_num':df.loc[df.index[index]].page_num, 'block_num':df.loc[df.index[index]].block_num, 'par_num':df.loc[df.index[index]].par_num, 'line_num':df.loc[df.index[index]].line_num, 'word_num':df.loc[df.index[index]].word_num, 'left':df.loc[df.index[index]].left, 'top':df.loc[df.index[index]].top, 'width':df.loc[df.index[index]].width, 'height':df.loc[df.index[index]].height, 'conf':df.loc[df.index[index]].conf, 'text':df.loc[df.index[index]].text}
                sdf = sdf.append(new_row, ignore_index=True)
                sindex += 1
            elif abs(sdf.loc[sdf.index[sindex-1]].top - df.loc[df.index[index]].top) < 15 and abs(sdf.loc[sdf.index[sindex-1]].left+sdf.loc[sdf.index[sindex-1]].width - df.loc[df.index[index]].left)<30:
                #if true, merge with current index row, otherwise add new row
                sdf.at[sdf.index[sindex-1],'text'] += " " + df.at[df.index[index],'text']
                space_length = int(df.at[df.index[index],'left']) - int(sdf.at[sdf.index[sindex-1],'left']) - int(sdf.at[sdf.index[sindex-1],'width'])
                sdf.at[sdf.index[sindex-1],'width'] =  int(sdf.at[sdf.index[sindex-1],'width']) + int(df.at[df.index[index],'width']) + space_length
            else:
                new_row = {'level': df.loc[df.index[index]].level, 'page_num':df.loc[df.index[index]].page_num, 'block_num':df.loc[df.index[index]].block_num, 'par_num':df.loc[df.index[index]].par_num, 'line_num':df.loc[df.index[index]].line_num, 'word_num':df.loc[df.index[index]].word_num, 'left':df.loc[df.index[index]].left, 'top':df.loc[df.index[index]].top, 'width':df.loc[df.index[index]].width, 'height':df.loc[df.index[index]].height, 'conf':df.loc[df.index[index]].conf, 'text':df.loc[df.index[index]].text}
                sdf = sdf.append(new_row, ignore_index=True)
                sindex += 1

            index += 1
        return sdf

    def label_finder(self, df, label, term, complete, x_delta, y_delta, constraint, sheet, row, column):
        
        for i in df.index:
            # can combine all the found results into one string
            for t in term:
                if complete=="NO":
                    if t in df.text[i]:
                        for j in df.index:
                            if int(df.left[i]) < int(df.left[j]):
                                label_x = int(df.left[i])+int(df.width[i])
                                value_x = int(df.left[j])    
                            else: #label is on the right and negative delta must be used (value of label is to the left of the label)
                                label_x = int(df.left[i])
                                value_x = int(df.left[j])+int(df.width[j])
                            label_y = int(df.top[i])
                            value_y = int(df.top[j])
                            #added last condition to make sure it only reads downwards
                            if min(label_x, label_x+x_delta) < value_x < max(label_x, label_x+x_delta) and abs(label_y-value_y) <y_delta and label_y < value_y+10:
                                if (':' in df.text[j]) == False and (t in df.text[j]) == False:
                                    if len(t) > 2 or t==df.text[i]: #this line just ensures you can't find PO in a word, might be better to make sure no letters are adjacent to t
                                        if constraint=="NUMBER":
                                            if re.match(r'^([\s\d]+)$', str(df.text[j])):
                                                append = str(sheet.cell(row,column).value)
                                                append += df.text[j]
                                                if append[0:4] == "None":
                                                    append = append[4:]
                                                sheet.cell(row,column).value = append
                                                print(label + ": " + str(df.text[j]) + ".........." + str(df.conf[j]) + "%")
                                                self.count += 1
                                        else:
                                            append = str(sheet.cell(row,column).value)
                                            append += df.text[j]+"\n"
                                            if append[0:4] == "None":
                                                append = append[4:]
                                            sheet.cell(row,column).value = append
                                            print(label + ": " + str(df.text[j]) + ".........." + str(df.conf[j]) + "%") 
                                            self.count += 1      
                else: 
                    if t in df.text[i]:
                        sheet.cell(row,column).value = label
                        print(label + ".........." + str(df.conf[i]) + "%")
                        self.count += 1
    
    def test(self):
        self.count += 1
        self.testText.set(self.count)
        self.label.pack()

    def setFormat(self, *args):
        print(self.format.get())
        self.format.set(self.format.get())

    def selectDirectory(self):
        self.directory = filedialog.askdirectory(initialdir="/", title='Select Folder')
        label = tk.Label(self.frame, text=self.directory+" [selected as directory]", bg="white")
        label.pack()

    def start(self):
        start_time = time.time()
        xls_file = pd.ExcelFile('labels.xlsx')
        #change the parse tab here to change template
        labels = xls_file.parse(self.format.get())
        
        #outputting to output.xlsx
        workbook = Workbook()
        sheet = workbook.active

        row = 2
        column = 1
        #outlining the output.xlsx
        for label in labels.index:
            sheet.cell(row,column).value = labels.LabelName[label]           
            row+= 1            
        
        row = 1
        column = 2

        with os.scandir(self.directory) as entries:
            for entry in entries:
                img = self.directory+'/'+entry.name
                img = cv2.imread(img)

                #PREPROCESS

                df = pytesseract.image_to_data(img, output_type='data.frame')
                df = df[df.conf != -1]
                sdf = self.spacify_dataframe(df)
                sheet.cell(row,column).value = entry.name
                row+=1 #next row after the entry names


                print("\n" + entry.name)

                for i in labels.index:
                    self.label_finder(sdf, labels.LabelName[i], labels.Term[i].split(','), labels.Complete[i],
                            labels.XDelta[i], labels.YDelta[i], labels.Constraint[i], sheet, row, column)
                    
                    row+= 1
                
                found = tk.Label(self.frame, text=entry.name + " found "+str(self.count)+" labels", bg="white")
                found.pack()
                self.count = 0
                
                #next file
                row = 1
                column += 1

                #empty dataframe    

        workbook.save(filename="output.xlsx")
        print("outputted to output.xlsx")

        print("--- %s seconds ---" % (time.time() - start_time))

        label = tk.Label(self.frame, text="Outputted to output.xlsx\n--- %s seconds ---" % (time.time() - start_time), bg="white")
        label.pack()
        

    def __init__(self, master=None):
        tk.Frame.__init__(self, master)

        self.winfo_toplevel().title("Auto William")

        self.directory = 'temp'
        self.format = 'ULINE'
        self.count = 0

        canvas = tk.Canvas(root, height=500, width=800, bg="#d88231")
        canvas.pack()

        self.frame = tk.Frame(root, bg="white")
        self.frame.place(relwidth=0.9, relheight=0.7, relx=0.05, rely=0.1)

        start = tk.Button(root, text="Start", padx=10, pady=5, fg="white", bg="#d88231", command=self.start)
        start.pack()

        selectDirectory = tk.Button(root, text="Select Directory", padx=10, pady=5, fg="white", bg="#d88231", command=self.selectDirectory)
        selectDirectory.pack()

        #test = tk.Button(root, text="TEST", padx=10, pady=5, fg="white", bg="#d88231", command=self.test)
        #test.pack()

        self.testText = tk.StringVar()
        self.testText.set("AUTOBILLING PROGRAM")
        self.label = tk.Label(self.frame, textvariable=self.testText, bg="white")
        self.label.pack()


        #Threshold Option
        #Preprocessing Options
        #Config Window for changing parameters in labels.xlsx
        #Swap output.xlsx rows and columns
        #show directory files
        #CHECKING OPTION
        #FIX LOGIC!!!
        
        OPTIONS = [
        "ULINE",
        "MCMASTER",
        "RAB",
        "SAFETYZONE",
        "PYLE",
        "OMG",
        "CARDINALHEALTH",
        "GENERAL",
        "WEBRESTAURANT",
        "TREK"
        ]

        self.format = StringVar()
        self.format.set(OPTIONS[0]) # default value

        w = OptionMenu(master, self.format, *OPTIONS)
        w.config(bg = "#d88231", fg="WHITE")
        w["menu"].config(bg="WHITE")
        w.pack()

        self.format.trace("w", self.setFormat)


if __name__ ==  '__main__':
    root = tk.Tk()
    MainApplication(root).pack(side="top", fill="both", expand=True)
    root.mainloop()